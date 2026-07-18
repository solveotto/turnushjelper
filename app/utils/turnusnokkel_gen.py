import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from app.utils import db_utils, df_utils


class TurnusnokkelGen():
    def __init__(self, turnus_name=None, turnus_set_id=None):
        self.turnus_set_id = turnus_set_id
        self.turnus_name = turnus_name
        self.sheet_name = 'Turnusnøkkel'
        self.start_row = 51
        self.start_col = 1

    def generate_single_turnus_nokkel(self):
        """
        Generate turnusnøkkel Excel file for a specific turnus
        
        Returns:
            dict: {'success': bool, 'filename': str, 'error': str}
        """
        try:
            if not self.turnus_name or not self.turnus_set_id:
                return {'success': False, 'error': 'Missing turnus name or turnus set ID'}
            
            # Get the turnus data for the specific turnus set
            df_manager = df_utils.DataframeManager(self.turnus_set_id)
            turnus_data = df_manager.turnus_data
            
            # Find the specific turnus data
            target_turnus_data = None
            for turnus_dict in turnus_data:
                if self.turnus_name in turnus_dict:
                    target_turnus_data = turnus_dict[self.turnus_name]
                    break
            
            if not target_turnus_data:
                return {'success': False, 'error': f'Turnus "{self.turnus_name}" not found in turnus set {self.turnus_set_id}'}
            
            # Get turnus set info for file paths
            turnus_set = db_utils.get_turnus_set_by_id(self.turnus_set_id)
            if not turnus_set:
                return {'success': False, 'error': f'Turnus set {self.turnus_set_id} not found'}
            
            year_identifier = turnus_set['year_identifier']
            
            # Define file paths based on turnus set
            # Get the project root directory
            project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            base_path = os.path.join(project_root, 'app', 'static', 'turnusfiler', year_identifier.lower())
            excel_template = os.path.join(base_path, f'turnusnøkkel_{year_identifier}_org.xlsx')
            
            # Debug info for error reporting
            debug_info = f"year_identifier={year_identifier}, base_path={base_path}, template={excel_template}"
            
            # Check if template exists
            if not os.path.exists(excel_template):
                # List available files in the directory for debugging
                available_files = []
                if os.path.exists(base_path):
                    available_files = os.listdir(base_path)
                
                error_msg = f'Template file not found: {excel_template}\n'
                error_msg += f'Debug info: {debug_info}\n'
                error_msg += f'Base path exists: {os.path.exists(base_path)}\n'
                error_msg += f'Available files in base path: {available_files}'
                return {'success': False, 'error': error_msg}
            
            # Load the existing Excel file
            workbook = load_workbook(excel_template)
            sheet = workbook.active
            
            # Access the specific sheet
            if self.sheet_name in workbook.sheetnames:
                sheet = workbook[self.sheet_name]
            else:
                return {'success': False, 'error': f"Sheet '{self.sheet_name}' not found in the Excel file"}
            
            for name in workbook.sheetnames:
                if name != self.sheet_name:
                    workbook[name].sheet_state = 'hidden'
            
            # Process the turnus data and insert into Excel
            # Skip non-dict entries (e.g. 'kl_timer', 'tj_timer' metadata strings)
            for uke_nr, ukedata in target_turnus_data.items():
                if not isinstance(ukedata, dict):
                    continue
                for dag_nr, dag_data in ukedata.items():
                    try:
                        start_value = dag_data['tid'][0] if dag_data['tid'] and len(dag_data['tid']) > 0 else ''
                    except (IndexError, KeyError, TypeError):
                        start_value = ''

                    try:
                        end_value = dag_data['tid'][1] if dag_data['tid'] and len(dag_data['tid']) > 1 else ''
                    except (IndexError, KeyError, TypeError):
                        end_value = ''
                    
                    # Create cell value
                    if start_value and end_value:
                        cell_value = f'{start_value} - {end_value}'
                    elif start_value:
                        cell_value = start_value
                    else:
                        cell_value = ''

                    # Set the value in the appropriate cell
                    col_letter = get_column_letter(self.start_col + int(dag_nr))
                    sheet[f"{col_letter}{self.start_row + int(uke_nr)}"] = cell_value

            # Reset scroll position so the sheet opens at the top
            sheet.sheet_view.topLeftCell = 'A1'

            # Generate filename (no need to save to disk)
            filename = f"Turnusnøkkel_{self.turnus_name}_{year_identifier}.xlsx"
            
            # Return the workbook object instead of saving to disk
            return {'success': True, 'filename': filename, 'workbook': workbook}
            
        except Exception as e:
            return {'success': False, 'error': f'Error generating turnusnøkkel: {str(e)}'}