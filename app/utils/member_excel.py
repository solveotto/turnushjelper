"""Parser for the NLF member list Excel file (Medlemsoppslag).

Required columns: ``Navn`` ("Etternavn, Fornavn") and ``Medlemsnr``.
Optional columns (new NLF export format): ``Innmeldt``, ``Født``, ``Tjenestested``.
Files without the optional columns still parse correctly (those fields return None).
"""
import logging
from datetime import date, datetime

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def _format_date(val) -> str | None:
    """Convert an openpyxl date/datetime value to DD.MM.YYYY string."""
    if val is None:
        return None
    if isinstance(val, (date, datetime)):
        return val.strftime("%d.%m.%Y")
    s = str(val).strip()
    if len(s) >= 10 and s[4] == "-":
        try:
            return datetime.strptime(s[:10], "%Y-%m-%d").strftime("%d.%m.%Y")
        except ValueError:
            pass
    return s or None


def parse_member_excel(path):
    """Parse a member-list xlsx into a list of member dicts.

    Each dict has keys: ``name``, ``medlemsnummer``, and optionally
    ``ans_dato``, ``fodt_dato``, ``stasjoneringssted`` (None when absent).

    Raises ``ValueError`` if the required header columns are missing.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        # Some export tools write a bogus dimension (e.g. "A1:A1"); read-only
        # mode trusts it and would truncate the sheet. Force a full read.
        ws.reset_dimensions()
        rows = ws.iter_rows(values_only=True)

        header = next(rows, None)
        if header is None:
            raise ValueError("Excel-fila er tom")
        col_index = {
            str(cell).strip().casefold(): i
            for i, cell in enumerate(header)
            if cell is not None
        }
        if "navn" not in col_index or "medlemsnr" not in col_index:
            raise ValueError(
                "Fant ikke kolonnene 'Navn' og 'Medlemsnr' i Excel-fila"
            )
        name_col = col_index["navn"]
        mnr_col = col_index["medlemsnr"]
        # Optional columns present in the richer NLF export format
        innmeldt_col = col_index.get("innmeldt")
        fodt_col = col_index.get("født")
        tjenestested_col = col_index.get("tjenestested")

        members = []
        for row in rows:
            name = row[name_col] if len(row) > name_col else None
            mnr = row[mnr_col] if len(row) > mnr_col else None
            if name is None and mnr is None:
                continue
            if isinstance(mnr, float) and mnr.is_integer():
                mnr = int(mnr)

            ans_dato = None
            if innmeldt_col is not None and len(row) > innmeldt_col:
                ans_dato = _format_date(row[innmeldt_col])

            fodt_dato = None
            if fodt_col is not None and len(row) > fodt_col:
                fodt_dato = _format_date(row[fodt_col])

            stasjoneringssted = None
            if tjenestested_col is not None and len(row) > tjenestested_col:
                raw = row[tjenestested_col]
                if raw:
                    # Strip company prefix: "VY OSLO" → "OSLO"
                    parts = str(raw).strip().split(None, 1)
                    stasjoneringssted = parts[1] if len(parts) > 1 else parts[0]

            members.append(
                {
                    "name": str(name).strip() if name is not None else "",
                    "medlemsnummer": str(mnr).strip() if mnr is not None else "",
                    "ans_dato": ans_dato,
                    "fodt_dato": fodt_dato,
                    "stasjoneringssted": stasjoneringssted,
                }
            )
        logger.info("parse_member_excel: %d rows parsed from %s", len(members), path)
        return members
    finally:
        wb.close()
